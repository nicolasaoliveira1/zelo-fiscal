"""Testes do relatorio de produtividade (spec 04, EXPORT-05).

Prova: agrega ExecucaoLote em emissoes/dia, taxa de sucesso por tipo e tempo
medio de lote; periodo vazio -> zeros; taxa nao divide por zero; nao inclui
custo de captcha; planilha abre valida.
"""
from datetime import timedelta

import pytest
from openpyxl import load_workbook

from app import db
from app.models import ExecucaoLote
from app.services import export_service
from app.utils import utcnow_naive


def _lote(tipo, *, sucesso=0, falhas=0, dias_atras=1, duracao_min=None, origem='manual'):
    iniciado = utcnow_naive() - timedelta(days=dias_atras)
    lote = ExecucaoLote(tipo=tipo, iniciado_em=iniciado, sucesso=sucesso,
                        falhas=falhas, origem=origem)
    if duracao_min is not None:
        lote.finalizado_em = iniciado + timedelta(minutes=duracao_min)
    return lote


@pytest.fixture()
def lotes(app):
    with app.app_context():
        db.create_all()
        db.session.add_all([
            # FGTS: 2 lotes, 8 sucessos, 2 falhas -> taxa 80%; tempos 10 e 20 -> media 15
            _lote('FGTS', sucesso=5, falhas=1, dias_atras=1, duracao_min=10),
            _lote('FGTS', sucesso=3, falhas=1, dias_atras=2, duracao_min=20),
            # Estadual RS: 1 lote, 4 sucessos, 0 falhas -> taxa 100% (sem div/0)
            _lote('Estadual RS', sucesso=4, falhas=0, dias_atras=1, duracao_min=None),
            # fora da janela (nao deve contar)
            _lote('Municipal', sucesso=99, falhas=99, dias_atras=90, duracao_min=5),
        ])
        db.session.commit()
        yield
        db.session.remove()
        db.drop_all()


def test_periodo_vazio_zera_tudo(app):
    with app.app_context():
        db.create_all()
        try:
            dados = export_service.coletar_produtividade(dias=30)
            assert dados['total_lotes'] == 0
            assert dados['total_emissoes'] == 0
            assert dados['tempo_medio_min'] is None
            assert dados['por_tipo'] == []
            assert dados['emissoes_por_dia'] == []
        finally:
            db.session.remove()
            db.drop_all()


def test_janela_exclui_lotes_antigos(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        # o lote Municipal de 90 dias atras fica de fora
        assert dados['total_lotes'] == 3
        assert {t['tipo'] for t in dados['por_tipo']} == {'FGTS', 'Estadual RS'}


def test_taxa_de_sucesso_por_tipo(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        por_tipo = {t['tipo']: t for t in dados['por_tipo']}
        assert por_tipo['FGTS']['sucesso'] == 8
        assert por_tipo['FGTS']['falhas'] == 2
        assert por_tipo['FGTS']['taxa'] == 80.0
        # 0 falhas nao divide por zero -> 100%
        assert por_tipo['Estadual RS']['taxa'] == 100.0


def test_tempo_medio_de_lote(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        por_tipo = {t['tipo']: t for t in dados['por_tipo']}
        assert por_tipo['FGTS']['tempo_medio_min'] == 15.0
        # Estadual sem finalizado_em -> sem media
        assert por_tipo['Estadual RS']['tempo_medio_min'] is None
        # media geral considera so os lotes finalizados (10 e 20)
        assert dados['tempo_medio_min'] == 15.0


def test_emissoes_por_dia(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        # dia -1: FGTS 5 + Estadual 4 = 9 ; dia -2: FGTS 3
        por_dia = {e['data']: e['emissoes'] for e in dados['emissoes_por_dia']}
        assert sorted(por_dia.values()) == [3, 9]
        assert dados['total_emissoes'] == 12


def test_nao_inclui_custo_de_captcha(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        chaves = set(dados) | {k for t in dados['por_tipo'] for k in t}
        assert not any('captcha' in c.lower() or 'custo' in c.lower() for c in chaves)


def test_planilha_produtividade_abre_valida(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        ws = load_workbook(export_service.gerar_planilha_produtividade(dados)).active
        assert ws.title == 'Produtividade'
        textos = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert 'Por tipo' in textos
        assert 'Emissões por dia' in textos


def test_segmenta_por_origem_sem_alterar_totais(app):
    """COV-04: manual vs agendador é um recorte adicional — os totais somam os dois
    e nenhum lote fica de fora."""
    with app.app_context():
        db.create_all()
        try:
            db.session.add_all([
                _lote('FGTS', sucesso=5, falhas=1, dias_atras=1, origem='manual'),
                _lote('FGTS', sucesso=3, falhas=0, dias_atras=2, origem='agendador'),
                _lote('Estadual RS', sucesso=4, falhas=0, dias_atras=1, origem='agendador'),
            ])
            db.session.commit()
            dados = export_service.coletar_produtividade(dias=30)

            assert dados['por_origem']['manual'] == {'lotes': 1, 'emissoes': 5}
            assert dados['por_origem']['agendador'] == {'lotes': 2, 'emissoes': 7}
            # o recorte por origem soma exatamente os totais gerais (nada escondido)
            somou_lotes = (dados['por_origem']['manual']['lotes']
                           + dados['por_origem']['agendador']['lotes'])
            somou_emissoes = (dados['por_origem']['manual']['emissoes']
                              + dados['por_origem']['agendador']['emissoes'])
            assert somou_lotes == dados['total_lotes'] == 3
            assert somou_emissoes == dados['total_emissoes'] == 12
        finally:
            db.session.remove()
            db.drop_all()


def test_origem_default_manual_para_registros_antigos(app):
    """Lote gravado sem origem (registro anterior à coluna) conta como manual."""
    with app.app_context():
        db.create_all()
        try:
            lote = ExecucaoLote(tipo='FGTS', iniciado_em=utcnow_naive(), sucesso=2)
            db.session.add(lote)
            db.session.commit()
            assert lote.origem == 'manual'  # default do modelo
            dados = export_service.coletar_produtividade(dias=30)
            assert dados['por_origem']['manual'] == {'lotes': 1, 'emissoes': 2}
            assert dados['por_origem']['agendador'] == {'lotes': 0, 'emissoes': 0}
        finally:
            db.session.remove()
            db.drop_all()


def test_registrar_execucao_lote_carimba_origem(app):
    """COV-04: o helper de registro (usado pela rota manual e pelo agendador)
    grava a origem. Default 'manual' (rota); 'agendador' quando o job passa."""
    from app.routes.lotes import _registrar_execucao_lote
    with app.app_context():
        db.create_all()
        try:
            _registrar_execucao_lote('FGTS', 'default', 3, 'exec-manual')
            _registrar_execucao_lote('FGTS', 'default', 2, 'exec-sched',
                                     origem='agendador')
            origens = {e.execution_id: e.origem for e in ExecucaoLote.query.all()}
            assert origens == {'exec-manual': 'manual', 'exec-sched': 'agendador'}
        finally:
            db.session.remove()
            db.drop_all()


# --- producao do agendador desde um corte (VGC-10/14/17) --------------------

def _horas_atras(horas):
    return utcnow_naive() - timedelta(hours=horas)


@pytest.fixture()
def corte(app):
    """Corte de 6h atras — a "passagem da noite" nos testes abaixo."""
    with app.app_context():
        db.create_all()
        yield _horas_atras(6)
        db.session.remove()
        db.drop_all()


def _add(*lotes):
    db.session.add_all(lotes)
    db.session.commit()


def _lote_em(tipo, horas_atras, **kw):
    iniciado = _horas_atras(horas_atras)
    lote = ExecucaoLote(tipo=tipo, iniciado_em=iniciado,
                        sucesso=kw.get('sucesso', 0), falhas=kw.get('falhas', 0),
                        origem=kw.get('origem', 'agendador'))
    if kw.get('duracao_min') is not None:
        lote.finalizado_em = iniciado + timedelta(minutes=kw['duracao_min'])
    return lote


def test_soma_emitidas_e_falhas_dos_lotes_do_agendador(app, corte):
    with app.app_context():
        _add(_lote_em('FGTS', 5, sucesso=20, falhas=2, duracao_min=30),
             _lote_em('Municipal', 4, sucesso=18, falhas=1, duracao_min=25))

        dados = export_service.coletar_producao_agendador(corte)

        assert dados['emitidas'] == 38
        assert dados['falhas'] == 3
        assert dados['lotes'] == 2


def test_lote_manual_na_janela_nao_entra(app, corte):
    """A faixa fala do que rodou SOZINHO; lote manual e o operador clicando."""
    with app.app_context():
        _add(_lote_em('FGTS', 5, sucesso=10, duracao_min=5),
             _lote_em('FGTS', 4, sucesso=99, origem='manual', duracao_min=5))

        assert export_service.coletar_producao_agendador(corte)['emitidas'] == 10


def test_lote_do_agendador_antes_do_corte_nao_entra(app, corte):
    with app.app_context():
        _add(_lote_em('FGTS', 7, sucesso=99, duracao_min=5))

        dados = export_service.coletar_producao_agendador(corte)

        assert dados['lotes'] == 0
        assert dados['emitidas'] == 0


def test_lote_sem_finalizado_em_conta_como_em_andamento(app, corte):
    """Numeros de lote em curso nao sao desfecho — quem exibe precisa saber."""
    with app.app_context():
        _add(_lote_em('FGTS', 5, sucesso=8, duracao_min=10),
             _lote_em('Municipal', 1, sucesso=3, duracao_min=None))

        dados = export_service.coletar_producao_agendador(corte)

        assert dados['em_andamento'] == 1
        assert dados['lotes'] == 2


def test_tipos_saem_sem_repetir_e_na_ordem_da_noite(app, corte):
    with app.app_context():
        _add(_lote_em('Municipal', 5, sucesso=1, duracao_min=5),
             _lote_em('FGTS', 3, sucesso=1, duracao_min=5),
             _lote_em('Municipal', 2, sucesso=1, duracao_min=5))

        assert export_service.coletar_producao_agendador(corte)['tipos'] == [
            'Municipal', 'FGTS']


def test_janela_sem_lote_devolve_zeros(app, corte):
    """Zeros aqui sao FATO da consulta. Interpretar "zero" como "nao rodou" ou
    como "nada havia" e decisao de quem exibe, nao desta funcao."""
    with app.app_context():
        dados = export_service.coletar_producao_agendador(corte)

        assert dados == {'lotes': 0, 'emitidas': 0, 'falhas': 0,
                         'em_andamento': 0, 'tipos': []}
